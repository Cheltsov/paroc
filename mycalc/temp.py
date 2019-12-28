import codecs, json
import os
import time

import schedule
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from paroc.settings import MEDIA_ROOT_W

f = codecs.open(MEDIA_ROOT_W + '\\insulations.txt', "r", "utf_8_sig")
insulations = f.read().split('\r\n')

filename = '../media/cal.xlsm'  # todo заменить txt файл на xlsm
wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

sheet = wb.get_sheet_by_name('Изоляция плоскостей')

# insulations[0] += (' - ' + str(sheet.cell(row=2, column=column_index_from_string('K')).value))
# print(insulations[0])
'''
_id = []
for i in range(40):
    row = i+2
    _ins = {}
    _i = str(i + 2)
    tol = []
    for row in sheet['K' + _i:'V' + _i]:
        for cell in row:
            if cell.value is not None:
                tol.append(str(cell.value))
    _ins['id'] = i
    _ins['name'] = insulations[i]
    _ins['tol'] = tol
    _id.append(_ins)

print(json.dumps(_id))
'''
'''
_id = []
for name in insulations:
    _ins = {}
    tol = []
    for row in range(40):
        print(row+2)
        if sheet.cell(row=row + 2, column=1).value == name:
            print('true')
            if sheet.cell(row=row + 2, column=2).value is not None:
                tol.append(str(sheet.cell(row=row + 2, column=2).value))
        print(tol)
        #_ins['id'] = row
        #_ins['name'] = name
        _ins['tol'] = tol
        _id.append(_ins)

print(json.dumps(_id))
'''
'''


path_to_dir = '../media/temp_files'

def remove_files():
    filelist = [f for f in os.listdir(path_to_dir) if f.endswith(".txt")]
    count = 0
    for f in filelist:
        os.remove(os.path.join(path_to_dir, f))
        count += 1
    print('Удалено ' + str(count) + ' файлов с расштрением .txt')


schedule.every(2).minutes.do(remove_files)

while True:
    schedule.run_pending()
    time.sleep(1)'''
'''
for i in range(43):
    row = i+2
    _i = str(i + 2)
    tol = []
    for row in sheet['K' + _i:'V' + _i]:
        for cell in row:
            if cell.value is not None:
                insulations += str(cell.value) + ','
    insulations[i] = insulations[i][:-1]
'''

f = codecs.open(MEDIA_ROOT_W + '\\insulations_plosk.txt', "r", "utf_8_sig")
insulations_plosk = f.read().split('\r\n')
    #_insulations_plosk = to_json('Изоляция плоскостей', insulations_plosk, 43)

ins = []
for i in range(43):
    by_id = {}
    _i = str(i + 2)
    tol = []
    for row in sheet['K' + _i:'V' + _i]:
        for cell in row:
            if cell.value is not None:
                tol.append(str(cell.value))
    by_id['id'] = i
    by_id['name'] = insulations_plosk[i]
    by_id['tol'] = tol
    ins.append(by_id)

print(json.dumps(ins))

