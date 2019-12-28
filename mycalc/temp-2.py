from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from django.template import TemplateDoesNotExist
from django.template.loader import get_template
from django.templatetags.static import static

from paroc.settings import MEDIA_ROOT_W
import mycalc.additional_functions as af
from mycalc.data import data_pipes as pipes
from mycalc.data import data_planes as planes
from mycalc.data import data_containers as containers
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from mycalc.additional_functions import to_json
from datetime import datetime
import codecs, json
import schedule
import time


# Create your views here.
def index(request):
    return render(request, 'mycalc/index.html')


def main(request):
    f = codecs.open(MEDIA_ROOT_W + '\\regions.txt', "r", "utf_8_sig")
    regions = f.read().split('\r\n')
    regions = sorted(regions)

    # f = codecs.open(MEDIA_ROOT_W + '\\insulations.txt', "r", "utf_8_sig")
    with open("media/insulations.json") as file1:
        insulations = json.load(file1)

    # f = codecs.open(MEDIA_ROOT_W + '\\insulations_plosk.txt', "r", "utf_8_sig")
    # insulations_plosk = f.read().split('\r\n')
    # _insulations_plosk = to_json('Изоляция плоскостей', insulations_plosk, 43)
    with open("media/insulations_plosk.json") as file2:
        insulations_plosk = json.load(file2)

    context = {
        'regions': regions,
        'insulations': insulations,
        'insulations_plosk': insulations_plosk
    }

    return render(request, 'mycalc/main.html', context)


def form(request):
    return render(request, 'mycalc/form.html')


def add(request):
    if request.method == 'POST':
        dirty_data = request.POST
        data_type = request.POST.getlist('type')

        # Первый символ верхнего регистра в сооветствии с названиями таблиц
        sheet_name = data_type[0][0].upper() + data_type[0][1:]

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v
        print(data)

        data_dicts = {'Trub': pipes.data,
                      'Plosk': planes.data,
                      'Emk': containers.data
                      }

        empty_dict = af.input_in_dict(data_dicts[sheet_name], data)

        filename = 'media/cal.xlsm'  # todo заменить txt файл на xlsm
        wb = load_workbook(filename=filename, read_only=False)

        sheet = wb.get_sheet_by_name(sheet_name)
        af.input_in_sheet(sheet, empty_dict)

        wb.save(filename='media/second-book.xlsx')

        return HttpResponse("post")
    else:
        return HttpResponse("no post")


def other_page_js(request, page):
    return redirect('/static/mycalc/js/' + page)


def other_page_form_js(request, page):
    return redirect('/static/mycalc/js/' + page)


def other_page_main_js(request, page):
    return redirect('/static/mycalc/js/' + page)


def add_trub(request):
    if request.method == 'POST':
        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'  # todo заменить txt файл на xlsm
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        sheet = wb.get_sheet_by_name('Trub')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Trub_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Trub_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Trub_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Trub_T_Sredi']

        sheet.cell(row=7, column=column_index_from_string('B')).value = data['L_Trub_WindSpeed']

        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Trub_Mater']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['CB_Trub_VneshPokr']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['cb_Usl_D']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Trub_Length']
        sheet.cell(row=14, column=column_index_from_string('B')).value = data['ChB_UsePoteri']
        sheet.cell(row=15, column=column_index_from_string('B')).value = data['CB_Trub_Krepezh']
        sheet.cell(row=16, column=column_index_from_string('B')).value = data['CB_Trub_Dir']
        sheet.cell(row=17, column=column_index_from_string('B')).value = data['ChB_Trub_Koltsa']
        sheet.cell(row=18, column=column_index_from_string('B')).value = data['L_Trub_Koltsa_Poteri']
        sheet.cell(row=19, column=column_index_from_string('B')).value = data['ChB_Trub_5000']
        sheet.cell(row=20, column=column_index_from_string('B')).value = data['L_Trub_D']
        sheet.cell(row=21, column=column_index_from_string('B')).value = data['L_Trub_WWidth']

        sheet.cell(row=23, column=column_index_from_string('B')).value = data['MP_Trub_Methods']

        sheet.cell(row=26, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Norm']

        sheet.cell(row=35, column=column_index_from_string('B')).value = data['CB_Trub_Iz_T']
        sheet.cell(row=36, column=column_index_from_string('B')).value = data['L_Trub_NosT2']
        sheet.cell(row=37, column=column_index_from_string('B')).value = data['L_Trub_Rashod_T']

        sheet.cell(row=47, column=column_index_from_string('B')).value = data['CB_Trub_Iz_MaxT']

        sheet.cell(row=56, column=column_index_from_string('B')).value = data['B_Trub_Iz_Cond']
        sheet.cell(row=57, column=column_index_from_string('B')).value = data['L_Hum']

        sheet.cell(row=66, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Peremerz']
        sheet.cell(row=67, column=column_index_from_string('B')).value = data['L_Trub_StopMove']

        sheet.cell(row=77, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Man']
        sheet.cell(row=78, column=column_index_from_string('B')).value = data['CB_Trub_Iz_W']
        sheet.cell(row=79, column=column_index_from_string('B')).value = data['LB_Trub_Iz']

        sheet.cell(row=91, column=column_index_from_string('B')).value = data['CB_Section']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        print(now.strftime("%d_%m_%Y %H_%M_%S"))
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Trub_Calc_Norm',
            '2': 'Trub_Calc_T',
            '3': 'Trub_Calc_MaxT',
            '4': 'Trub_Calc_Cond',
            '5': 'Trub_Calc_Permerz',
            '6': 'Trub_Calc_Man',
        }

        flag = data['flat_isol']

        result = macro_run(flags[flag], temp_cal)  # запихать в json
        print(result)

        return HttpResponse(json.dumps(result))
    else:
        return HttpResponse('no post')


def add_plosk(request):
    if request.method == 'POST':

        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'  # todo заменить txt файл на xlsm
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        sheet = wb.get_sheet_by_name('Plosk')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Plosk_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Plosk_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Plosk_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Plosk_T_Sredi']
        sheet.cell(row=6, column=column_index_from_string('B')).value = data['L_Plosk_WindSpeed']

        sheet.cell(row=8, column=column_index_from_string('B')).value = data['CB_Plosk_Mater']
        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Plosk_VneshPokr']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['L_Plosk_Length']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['L_Plosk_WWidth']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Plosk_Width']
        sheet.cell(row=13, column=column_index_from_string('B')).value = data['ChB_Plosk_5000']

        sheet.cell(row=15, column=column_index_from_string('B')).value = data['MP_Plosk_Methods']

        sheet.cell(row=18, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Norm']

        sheet.cell(row=27, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_MaxT']

        sheet.cell(row=36, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Cond']

        sheet.cell(row=46, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Man']
        sheet.cell(row=47, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_W']
        sheet.cell(row=48, column=column_index_from_string('B')).value = data['LB_Plosk_Iz']

        sheet.cell(row=60, column=column_index_from_string('B')).value = data['CB_Plosk_Section']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Plosk_Calc_Norm',
            '3': 'Plosk_Calc_MaxT',
            '4': 'Plosk_Calc_Cond',
            '6': 'Plosk_Calc_Man',
        }

        flag = data['flat_isol']

        result = macro_run(flags[flag], temp_cal)
        print(result)

        return HttpResponse(json.dumps(result))
    else:
        return HttpResponse('no post')


def add_emk(request):
    if request.method == 'POST':

        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'  # todo заменить txt файл на xlsm
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        sheet = wb.get_sheet_by_name('Plosk')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Emk_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Emk_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Emk_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Emk_T_Sredi']
        sheet.cell(row=6, column=column_index_from_string('B')).value = data['L_Emk_WindSpeed']

        sheet.cell(row=8, column=column_index_from_string('B')).value = data['CB_Emk_Mater']
        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Emk_VneshPokr']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['ChB_Emk_5000']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['L_Emk_Height']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Emk_Diam']
        sheet.cell(row=13, column=column_index_from_string('B')).value = data['ChB_UseDnishe']
        sheet.cell(row=14, column=column_index_from_string('B')).value = data['L_Emk_WWidth']
        sheet.cell(row=15, column=column_index_from_string('B')).value = data['L_Emk_WPlotn']
        sheet.cell(row=16, column=column_index_from_string('B')).value = data['L_Emk_WC']

        sheet.cell(row=18, column=column_index_from_string('B')).value = data['MP_Emk_Methods']

        sheet.cell(row=21, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Norm']

        sheet.cell(row=31, column=column_index_from_string('B')).value = data['CB_Emk_Iz_T']
        sheet.cell(row=32, column=column_index_from_string('B')).value = data['L_Emk_NosT2']
        sheet.cell(row=33, column=column_index_from_string('B')).value = data['L_Emk_THran']

        sheet.cell(row=43, column=column_index_from_string('B')).value = data['CB_Emk_Iz_MaxT']

        sheet.cell(row=53, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Cond']
        sheet.cell(row=54, column=column_index_from_string('B')).value = data['L_Emk_Hum']

        sheet.cell(row=64, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Man']
        sheet.cell(row=65, column=column_index_from_string('B')).value = data['CB_Emk_Iz_W']
        sheet.cell(row=66, column=column_index_from_string('B')).value = data['LB_Emk_Iz']

        sheet.cell(row=79, column=column_index_from_string('B')).value = data['CB_Emk']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Emk_Calc_Norm',
            '2': 'Emk_Calc_T',
            '3': 'Emk_Calc_MaxT',
            '4': 'Emk_Calc_Cond',
            '6': 'Emk_Calc_Man',
        }

        flag = data['flat_isol']

        result = macro_run(flags[flag], temp_cal)
        print(result)

        return HttpResponse(json.dumps(result))
    else:
        return HttpResponse('no post')


def macro_run(macros_name, cal_empty_copy):
    from datetime import datetime
    import win32com.client as wincl
    import os
    from os.path import join, abspath
    now = datetime.now()

    filename = cal_empty_copy
    wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)
    sheet = wb.get_sheet_by_name('communication')
    result_file = "result_" + now.strftime("%d_%m_%Y %H_%M_%S")
    for_checking = filename
    sheet.cell(row=1, column=column_index_from_string('B')).value = result_file

    wb.save(filename)

    data_path = join(',', filename)
    data_path = abspath(data_path)

    excel_macro = wincl.DispatchEx("Excel.application")
    excel_path = os.path.expanduser(data_path)

    if os.path.exists(excel_path):
        workbook = excel_macro.Workbooks.Open(Filename=excel_path, ReadOnly=1)
        print('Run macros= ' + macros_name)
        excel_macro.Application.Run(macros_name)
        workbook.Save()
        excel_macro.Application.Quit()
        del excel_macro

    result_file = 'media/temp_files/' + result_file + '.xlsx'
    wb = load_workbook(filename=result_file, data_only=True, read_only=False, keep_vba=True)
    sheet = wb.get_sheet_by_name('Протокол')

    result_list = {'accurate-calculation': None,
                   'heat-loss': sheet.cell(row=28, column=column_index_from_string('G')).value,
                   'recommended-thickness': sheet.cell(row=29, column=column_index_from_string('G')).value,
                   'permissible-temperature': sheet.cell(row=30, column=column_index_from_string('G')).value,
                   'surface-temperature': sheet.cell(row=31, column=column_index_from_string('G')).value}

    path_to_dir = 'media/temp_files'

    if for_checking[18:20] == "02":
        af.remove_files(path_to_dir)

    return result_list