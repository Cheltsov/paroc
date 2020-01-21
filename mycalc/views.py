from django.shortcuts import render, redirect
from django.http import HttpResponse

from paroc.settings import MEDIA_ROOT_W
import mycalc.additional_functions as af
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from django import template
import codecs
import shutil
import json

import logging

logger = logging.getLogger(__name__)

from datetime import datetime

now = datetime.now()


def index(request):
    logger.info(request)
    return render(request, 'mycalc/index.html')


def main(request):
    logger.info(request)
    f = codecs.open(MEDIA_ROOT_W + '\\regions.txt', "r", "utf_8_sig")
    regions = f.read().split('\r\n')
    regions = sorted(regions)

    with open("media/insulations.json") as f1:
        insulations = json.load(f1)

    with open("media/insulations_plosk.json") as f2:
        insulations_plosk = json.load(f2)

    context = {
        'regions': regions,
        'insulations': insulations["insulations"],
        'insulations_plosk': insulations_plosk["insulations_plosk"]
    }
    return render(request, 'mycalc/main.html', context)


def form(request):
    logger.info(request)
    return render(request, 'mycalc/form.html')


def other_page_js(request, page):
    logger.info(request)
    return redirect('/static/mycalc/js/' + page)


def other_page_form_js(request, page):
    logger.info(request)
    return redirect('/static/mycalc/js/' + page)


def other_page_main_js(request, page):
    logger.info(request)
    return redirect('/static/mycalc/js/' + page)


def add(request):
    pass


def add_trub(request):
    logger.info(request)
    if request.method == 'POST':
        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        error = ''
        err_sheet = wb.get_sheet_by_name('communication')
        error = err_sheet.cell(row=2, column=column_index_from_string('B')).value

        sheet = wb.get_sheet_by_name('Trub')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Trub_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Trub_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Trub_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Trub_T_Sredi']

        sheet.cell(row=7, column=column_index_from_string('B')).value = data['L_Trub_WindSpeed']

        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Trub_Mater']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['CB_Trub_VneshPokr']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['cb_Usl_D']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Trub_Length']
        sheet.cell(row=14, column=column_index_from_string('B')).value = data['ChB_UsePoteri']
        sheet.cell(row=15, column=column_index_from_string('B')).value = data['CB_Trub_Krepezh']
        sheet.cell(row=16, column=column_index_from_string('B')).value = data['CB_Trub_Dir']
        sheet.cell(row=17, column=column_index_from_string('B')).value = data['ChB_Trub_Koltsa']
        sheet.cell(row=18, column=column_index_from_string('B')).value = data['L_Trub_Koltsa_Poteri']
        sheet.cell(row=19, column=column_index_from_string('B')).value = data['ChB_Trub_5000']
        sheet.cell(row=20, column=column_index_from_string('B')).value = data['L_Trub_D']
        sheet.cell(row=21, column=column_index_from_string('B')).value = data['L_Trub_WWidth']

        sheet.cell(row=23, column=column_index_from_string('B')).value = data['MP_Trub_Methods']

        sheet.cell(row=26, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Norm']

        sheet.cell(row=35, column=column_index_from_string('B')).value = data['CB_Trub_Iz_T']
        sheet.cell(row=36, column=column_index_from_string('B')).value = data['L_Trub_NosT2']
        sheet.cell(row=37, column=column_index_from_string('B')).value = data['L_Trub_Rashod_T']

        sheet.cell(row=47, column=column_index_from_string('B')).value = data['CB_Trub_Iz_MaxT']

        sheet.cell(row=56, column=column_index_from_string('B')).value = data['B_Trub_Iz_Cond']
        sheet.cell(row=57, column=column_index_from_string('B')).value = data['L_Hum']

        sheet.cell(row=66, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Peremerz']
        sheet.cell(row=67, column=column_index_from_string('B')).value = data['L_Trub_StopMove']

        sheet.cell(row=77, column=column_index_from_string('B')).value = data['CB_Trub_Iz_Man']
        sheet.cell(row=78, column=column_index_from_string('B')).value = data['CB_Trub_Iz_W']

        sheet.cell(row=91, column=column_index_from_string('B')).value = data['CB_Section']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Trub_Calc_Norm',
            '2': 'Trub_Calc_T',
            '3': 'Trub_Calc_MaxT',
            '4': 'Trub_Calc_Cond',
            '5': 'Trub_Calc_Permerz',
            '6': 'Trub_Calc_Man',
        }

        Trub_Calc_Norm = {'4': 29, '1': 30, '2': 31, '5': 32, '3': 1, '6': 1, '7': 1, '8': 1}
        Trub_Calc_T = {'1': 32, '8': 33, '2': 34, '3': 35, '4': 36, '5': 37, '7': 1, '6': 1}
        Trub_Calc_MaxT = {'1': 29, '2': 30, '3': 31, '4': 32, '5': 33, '6': 1, '7': 1, '8': 1}
        Trub_Calc_Cond = {'1': 31, '2': 32, '3': 33, '4': 34, '5': 35, '6': 1, '7': 1, '8': 1}
        Trub_Calc_Permerz = {'1': 30, '4': 31, '5': 32, '2': 33, '3': 34, '6': 1, '7': 1, '8': 1}
        Trub_Calc_Man = {'4': 33, '5': 34, '2': 35, '3': 36, '6': 37, '1': 1, '7': 1, '8': 1}

        flags_data = {
            '1': Trub_Calc_Norm,
            '2': Trub_Calc_T,
            '3': Trub_Calc_MaxT,
            '4': Trub_Calc_Cond,
            '5': Trub_Calc_Permerz,
            '6': Trub_Calc_Man,
        }

        flag = data['flat_isol']

        result = macro_run(flags[flag], flags_data[flag], temp_cal, 0, error)
        print(result)

        logger.debug(HttpResponse)
        return HttpResponse(json.dumps(result))
    else:
        logger.error(HttpResponse)
        return HttpResponse('no post')


def add_plosk(request):
    logger.info(request)
    if request.method == 'POST':

        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        error = ''
        err_sheet = wb.get_sheet_by_name('communication')
        error = err_sheet.cell(row=2, column=column_index_from_string('B')).value

        sheet = wb.get_sheet_by_name('Plosk')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Plosk_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Plosk_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Plosk_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Plosk_T_Sredi']
        sheet.cell(row=6, column=column_index_from_string('B')).value = data['L_Plosk_WindSpeed']

        sheet.cell(row=8, column=column_index_from_string('B')).value = data['CB_Plosk_Mater']
        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Plosk_VneshPokr']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['L_Plosk_Length']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['L_Plosk_WWidth']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Plosk_Width']
        sheet.cell(row=13, column=column_index_from_string('B')).value = data['ChB_Plosk_5000']

        sheet.cell(row=15, column=column_index_from_string('B')).value = data['MP_Plosk_Methods']

        sheet.cell(row=18, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Norm']

        sheet.cell(row=27, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_MaxT']

        sheet.cell(row=36, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Cond']

        sheet.cell(row=46, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_Man']
        sheet.cell(row=47, column=column_index_from_string('B')).value = data['CB_Plosk_Iz_W']
        sheet.cell(row=48, column=column_index_from_string('B')).value = data['LB_Plosk_Iz']

        sheet.cell(row=60, column=column_index_from_string('B')).value = data['CB_Plosk_Section']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Plosk_Calc_Norm',
            '3': 'Plosk_Calc_MaxT',
            '4': 'Plosk_Calc_Cond',
            '6': 'Plosk_Calc_Man',
        }

        Plosk_Calc_Norm = {'4': 24, '1': 25, '2': 26, '3': 27, '5': 1, '6': 1, '7': 1, '8': 1}
        Plosk_Calc_MaxT = {'1': 24, '2': 25, '3': 26, '4': 27, '5': 28, '6': 1, '7': 1, '8': 1}
        Plosk_Calc_Cond = {'1': 1, '2': 1, '3': 1, '4': 1, '5': 1, '6': 1, '7': 1, '8': 1}  # todo
        Plosk_Calc_Man = {'1': 1, '2': 1, '3': 1, '4': 1, '5': 1, '6': 1, '7': 1, '8': 1}  # todo

        flags_data = {
            '1': Plosk_Calc_Norm,
            '3': Plosk_Calc_MaxT,
            '4': Plosk_Calc_Cond,
            '6': Plosk_Calc_Man,
        }

        flag = data['flat_isol']
        print(data['flat_isol'])

        result = macro_run(flags[flag], flags_data[flag], temp_cal, 1, error)
        print(result)

        logger.debug(HttpResponse)
        return HttpResponse(json.dumps(result))
    else:
        logger.error(HttpResponse)
        return HttpResponse('no post')


def add_emk(request):
    logger.info(request)
    if request.method == 'POST':

        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        filename = 'media/cal.xlsm'
        wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)

        error = ''
        err_sheet = wb.get_sheet_by_name('communication')
        error = err_sheet.cell(row=2, column=column_index_from_string('B')).value

        sheet = wb.get_sheet_by_name('Plosk')

        sheet.cell(row=2, column=column_index_from_string('B')).value = data['CB_Emk_Region']
        sheet.cell(row=3, column=column_index_from_string('B')).value = data['CB_Emk_Sreda']
        sheet.cell(row=4, column=column_index_from_string('B')).value = data['L_Emk_NosT']
        sheet.cell(row=5, column=column_index_from_string('B')).value = data['L_Emk_T_Sredi']
        sheet.cell(row=6, column=column_index_from_string('B')).value = data['L_Emk_WindSpeed']

        sheet.cell(row=8, column=column_index_from_string('B')).value = data['CB_Emk_Mater']
        sheet.cell(row=9, column=column_index_from_string('B')).value = data['CB_Emk_VneshPokr']
        sheet.cell(row=10, column=column_index_from_string('B')).value = data['ChB_Emk_5000']
        sheet.cell(row=11, column=column_index_from_string('B')).value = data['L_Emk_Height']
        sheet.cell(row=12, column=column_index_from_string('B')).value = data['L_Emk_Diam']
        sheet.cell(row=13, column=column_index_from_string('B')).value = data['ChB_UseDnishe']
        sheet.cell(row=14, column=column_index_from_string('B')).value = data['L_Emk_WWidth']
        sheet.cell(row=15, column=column_index_from_string('B')).value = data['L_Emk_WPlotn']
        sheet.cell(row=16, column=column_index_from_string('B')).value = data['L_Emk_WC']

        sheet.cell(row=18, column=column_index_from_string('B')).value = data['MP_Emk_Methods']

        sheet.cell(row=21, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Norm']

        sheet.cell(row=31, column=column_index_from_string('B')).value = data['CB_Emk_Iz_T']
        sheet.cell(row=32, column=column_index_from_string('B')).value = data['L_Emk_NosT2']
        sheet.cell(row=33, column=column_index_from_string('B')).value = data['L_Emk_THran']

        sheet.cell(row=43, column=column_index_from_string('B')).value = data['CB_Emk_Iz_MaxT']

        sheet.cell(row=53, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Cond']
        sheet.cell(row=54, column=column_index_from_string('B')).value = data['L_Emk_Hum']

        sheet.cell(row=64, column=column_index_from_string('B')).value = data['CB_Emk_Iz_Man']
        sheet.cell(row=65, column=column_index_from_string('B')).value = data['CB_Emk_Iz_W']
        sheet.cell(row=66, column=column_index_from_string('B')).value = data['LB_Emk_Iz']

        sheet.cell(row=79, column=column_index_from_string('B')).value = data['CB_Emk']

        now = datetime.now()
        temp_cal = 'media/temp_files/cal' + now.strftime("%d_%m_%Y %H_%M_%S") + '.xlsm'
        wb.save(temp_cal)
        wb.close()

        flags = {
            '1': 'Emk_Calc_Norm',
            '2': 'Emk_Calc_T',
            '3': 'Emk_Calc_MaxT',
            '4': 'Emk_Calc_Cond',
            '6': 'Emk_Calc_Man',
        }

        Emk_Calc_Norm = {'4': 26, '5': 27, '1': 28, '2': 29, '3': 30, '6': 1, '7': 1, '8': 1}
        Emk_Calc_T = {'1': 29, '2': 30, '3': 31, '4': 32, '5': 33, '6': 1, '7': 34, '8': 1}
        Emk_Calc_MaxT = {'1': 26, '2': 27, '3': 28, '4': 29, '5': 30, '6': 1, '7': 31, '8': 1}
        Emk_Calc_Cond = {'1': 1, '2': 1, '3': 1, '4': 1, '5': 1, '6': 1, '7': 1, '8': 1}
        Emk_Calc_Man = {'4': 30, '5': 31, '7': 32, '2': 33, '3': 34, '6': 35, '1': 1, '8': 1}

        flags_data = {
            '1': Emk_Calc_Norm,
            '2': Emk_Calc_T,
            '3': Emk_Calc_MaxT,
            '4': Emk_Calc_Cond,
            '6': Emk_Calc_Man,
        }

        flag = data['flat_isol']

        result = macro_run(flags[flag], flags_data[flag], temp_cal, 2, error)
        print(result)

        logger.debug(HttpResponse)
        return HttpResponse(json.dumps(result))
    else:
        logger.error(HttpResponse)
        return HttpResponse('no post')


def macro_run(macros_name, macro_data, cal_empty_copy, first_macro_name, error):
    import win32com.client as wincl
    import os, pythoncom
    from os.path import join, abspath

    print(macro_data)

    print('error: ', error)

    for_checking = filename = cal_empty_copy
    wb = load_workbook(filename=filename, data_only=True, read_only=False, keep_vba=True)
    sheet = wb.get_sheet_by_name('communication')
    result_file = "result_" + now.strftime("%d_%m_%Y %H_%M_%S")
    sheet.cell(row=1, column=column_index_from_string('B')).value = result_file

    wb.save(filename)

    data_path = join('', filename)
    data_path = abspath(data_path)

    pythoncom.CoInitialize()
    try:
        excel_macro = wincl.DispatchEx("Excel.application")
        excel_path = os.path.expanduser(data_path)

        workbook = excel_macro.Workbooks.Open(Filename=excel_path, ReadOnly=1)
        print('Run macro = ' + macros_name)

        fill_data_macros = ['Fill_Data_Form_Trub', 'Fill_Data_Form_Plosk', 'Fill_Data_Form_Emk']
        excel_macro.Application.Run(fill_data_macros[first_macro_name])
        excel_macro.Application.Run(macros_name)

        workbook.Save()
        excel_macro.Application.Quit()
        del excel_macro
    finally:
        pythoncom.CoUninitialize()

    result_file = 'media/temp_files/' + result_file + '.xlsx'

    wb = load_workbook(filename=result_file, data_only=True, read_only=False, keep_vba=True)

    sheet = wb.get_sheet_by_name('Протокол')

    result_list = {
        'recommended-thickness': sheet.cell(row=macro_data['1'], column=column_index_from_string('G')).value,
        # 1 Рекомендуемая толщина выбанной изоляции
        'permissible-temperature': sheet.cell(row=macro_data['2'], column=column_index_from_string('G')).value,
        # 2 Максимально допустимая температура поверхности изоляции
        'surface-temperature': sheet.cell(row=macro_data['3'], column=column_index_from_string('G')).value,
        # 3 Расчётная температура поверхности изоляции
        'heat-loss-SP': sheet.cell(row=macro_data['4'], column=column_index_from_string('G')).value,
        # 4 Тепловые потери согласно нормам СП
        'estimated-heat-loss': sheet.cell(row=macro_data['5'], column=column_index_from_string('G')).value,
        # 5 Расчётные тепловые потери
        'layer-temperature': sheet.cell(row=macro_data['6'], column=column_index_from_string('G')).value,
        # 6 Температура на границе слоя
        'total-estimated-heat-loss': sheet.cell(row=macro_data['7'], column=column_index_from_string('G')).value,
        # 7 Полные расчётные тепловые потери
        'final-temperature': sheet.cell(row=macro_data['8'], column=column_index_from_string('G')).value,
        # 8 Конечная температура теплоносителя

        'error': error,

        'type': macros_name[0],
        'macro-name': macros_name
    }

    path_to_dir = 'media/temp_files'

    if os.path.exists('media/Ведомость.xlsx'):
        os.remove(os.path.join('media/Ведомость.xlsx'))
    shutil.copyfile(result_file, 'media/Ведомость.xlsx')

    if for_checking[18:20] == "02":
        af.remove_files(path_to_dir)

    return result_list


register = template.Library()
temp = "nothing"


def make_result_file(request):
    logger.info(request)
    if request.method == 'POST':
        print('make_result_file')
        dirty_data = request.POST

        data = {}
        for k, v in dirty_data.items():
            data[k[5:-1]] = v

        word = data["name"] + data["email"] + data["phone"] + now.strftime("%d_%m_%Y %H_%M_%S")
        name = af.hash_word(word)
        path = "media/download/" + name + ".xlsx"

        shutil.copyfile('media/Ведомость.xlsx', path)

        path = '/'+path
        print("data ----", path)

        logger.debug(HttpResponse)
        return HttpResponse(path)
    else:
        logger.error(HttpResponse)
        return HttpResponse('no post')


'recommended-thickness'
'permissible-temperature'
'surface-temperature'
'heat-loss-SP'
'estimated-heat-loss'
'layer-temperature'
'total-estimated-heat-loss'

# Рекомендуемая толщина выбанной изоляции
# Максимально допустимая температура поверхности изоляции
# Расчётная температура поверхности изоляции
# Тепловые потери согласно нормам СП
# Расчётные тепловые потери
# Температура на границе слоя
# Полные расчётные тепловые потери
