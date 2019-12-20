from get_data_from_js import data_pipes as pipes
from get_data_from_js import data_planes as planes
from get_data_from_js import data_containers as containers
from openpyxl import load_workbook

filename = 'Калькулятор Парок ТИ 19_12_17.xlsm'
wb = load_workbook(filename=filename, read_only=False)

sheet_names = ['Trub', 'Plosk', 'Emk']
data_dicts = [pipes.data, planes.data, containers.data]


def input_in_sheet(sheet_name, data):
    sheet = wb.get_sheet_by_name(sheet_name)

    for key,value in data.items():
        for row in range(len(data)):
            if sheet.cell(row=row + 2, column=1).value == key:
                sheet.cell(row=row + 2, column=2, value= value)
                break


for i in range(len(sheet_names)):
    input_in_sheet(sheet_names[i], data_dicts[i])

wb.save(filename='second-book.xlsx')